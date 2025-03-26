from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import os
import json
import uuid
import re
import pandas as pd
from werkzeug.utils import secure_filename
import excel_analyzer  # 导入现有的分析脚本
import openpyxl  # 直接导入openpyxl，避免通过excel_analyzer调用
import xlwings as xw  # 导入xlwings用于Excel计算

# 更新说明：
# 2023年更新 - 放弃使用formulas库进行计算，改为使用xlwings直接调用Excel进行计算
# 通过xlwings，我们可以将前端输入的参数传入Excel，让Excel进行计算，
# 然后读取计算结果返回给前端，实现更准确的计算和更好的兼容性。

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# 确保上传目录存在
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 辅助函数：检查文件扩展名
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# 首页路由 - 显示上传表单
@app.route('/')
def index():
    return render_template('index.html')

# 处理文件上传
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件部分'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '未选择文件'}), 400
    
    if file and allowed_file(file.filename):
        # 创建唯一的会话ID
        session_id = str(uuid.uuid4())
        session['session_id'] = session_id
        
        # 安全地保存文件
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_id}_{filename}")
        file.save(file_path)
        
        try:
            # 调用excel_analyzer分析文件 - 这会生成优化后的Excel文件
            all_params = excel_analyzer.analyze_excel(file_path)
            
            # 获取优化后的Excel文件路径
            optimized_file_path = os.path.splitext(file_path)[0] + "_optimized.xlsx"
            
            # 检查优化后的文件是否存在
            if not os.path.exists(optimized_file_path):
                print(f"警告: 优化后的Excel文件不存在: {optimized_file_path}")
                print("将使用原始文件继续处理")
                optimized_file_path = file_path
            else:
                print(f"使用优化后的Excel文件: {optimized_file_path}")
            
            # 保存优化后的文件路径到会话
            session['file_path'] = optimized_file_path
            session['original_file_path'] = file_path
            session['analyzed'] = True
            
            return jsonify({'success': True, 'redirect': url_for('visualize')})
        except Exception as e:
            return jsonify({'error': f'分析文件时出错: {str(e)}'}), 500
    
    return jsonify({'error': '不支持的文件类型'}), 400

# 可视化页面路由
@app.route('/visualize')
def visualize():
    if not session.get('analyzed', False):
        return redirect(url_for('index'))
    
    return render_template('visualization.html')

# API: 获取所有参数及其分类
@app.route('/api/parameters')
def get_parameters():
    if not session.get('file_path'):
        return jsonify({'error': '找不到已分析的文件'}), 404
    
    try:
        file_path = session['file_path']
        print(f"正在加载优化后的文件: {file_path}")
        
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            # 尝试回退到原始文件
            if session.get('original_file_path') and os.path.exists(session['original_file_path']):
                file_path = session['original_file_path']
                print(f"回退到原始文件: {file_path}")
            else:
                return jsonify({'error': f'文件不存在: {file_path}'}), 404
        
        try:
            # 重新加载Excel工作簿，确保数据是最新的
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            
            # 打印工作表信息进行调试
            print(f"工作表列表: {wb.sheetnames}")
            
            # 收集参数信息
            all_params, formula_dependencies = excel_analyzer.collect_params_and_dependencies(wb_formulas, wb, {})
            
            # 检查结果
            if not all_params:
                print("没有找到任何参数")
                return jsonify({'error': '没有找到任何参数，请检查Excel文件格式是否正确'}), 400
            
            print(f"找到 {len(all_params)} 个参数")
            
            # 对参数进行分类
            input_params, output_params, intermediate_params, independent_params = excel_analyzer.categorize_parameters(all_params, formula_dependencies)
            
            # 将所有参数中的set类型转换为list类型，以便于JSON序列化
            def convert_sets_to_lists(params_dict):
                for param_id, param_info in params_dict.items():
                    if "依赖" in param_info and isinstance(param_info["依赖"], set):
                        param_info["依赖"] = list(param_info["依赖"])
                    if "依赖描述" in param_info and isinstance(param_info["依赖描述"], set):
                        param_info["依赖描述"] = list(param_info["依赖描述"])
                return params_dict
            
            # 转换所有参数中的set为list
            all_params = convert_sets_to_lists(all_params)
            
            # 确保参数结构一致性
            def normalize_param_info(param_info):
                # 确保基本属性存在
                required_fields = ['名称', '标识符', '单位', '值', '公式', '公式描述', '依赖', '依赖描述', '有循环依赖']
                for field in required_fields:
                    if field not in param_info:
                        if field in ['依赖', '依赖描述']:
                            param_info[field] = []
                        elif field in ['公式', '公式描述', '单位']:
                            param_info[field] = ''
                        elif field == '值':
                            param_info[field] = 0
                        elif field == '有循环依赖':
                            param_info[field] = False
                        else:
                            param_info[field] = param_info.get('名称', '')
                return param_info
            
            # 标准化每个参数的信息
            for param_id in all_params:
                all_params[param_id] = normalize_param_info(all_params[param_id])
            
            # 组织数据以便前端使用
            parameters = {
                'input_params': [all_params[param_id] for param_id in input_params],
                'output_params': [all_params[param_id] for param_id in output_params],
                'intermediate_params': [all_params[param_id] for param_id in intermediate_params],
                'independent_params': [all_params[param_id] for param_id in independent_params]
            }
            
            return jsonify(parameters)
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print(f"无效的Excel文件: {str(e)}")
            return jsonify({'error': f'无效的Excel文件格式，请确保文件可以正常在Excel中打开: {str(e)}'}), 400
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"获取参数时出错: {str(e)}")
        print(f"错误详情: {error_details}")
        return jsonify({'error': f'获取参数时出错: {str(e)}'}), 500

# API: 获取依赖关系
@app.route('/api/dependencies')
def get_dependencies():
    if not session.get('file_path'):
        return jsonify({'error': '找不到已分析的文件'}), 404
    
    try:
        file_path = session['file_path']
        print(f"正在加载依赖关系的文件: {file_path}")
        
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            # 尝试回退到原始文件
            if session.get('original_file_path') and os.path.exists(session['original_file_path']):
                file_path = session['original_file_path']
                print(f"回退到原始文件: {file_path}")
            else:
                return jsonify({'error': f'文件不存在: {file_path}'}), 404
        
        # 重新加载Excel工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        
        # 收集参数信息和依赖关系
        all_params, formula_dependencies = excel_analyzer.collect_params_and_dependencies(wb_formulas, wb, {})
        
        # 检查数据结构
        if not formula_dependencies or not isinstance(formula_dependencies, dict):
            print("警告: 依赖关系数据不是预期的字典格式")
            formula_dependencies = {}
        
        # 格式化依赖关系
        dependencies = []
        for param_id, deps in formula_dependencies.items():
            # 确保deps是列表或集合类型
            if not deps:
                continue
                
            deps_list = list(deps) if isinstance(deps, set) else deps
            
            source_name = "未知参数"
            if param_id in all_params:
                source_name = all_params[param_id].get('名称', param_id)
            
            for dep_id in deps_list:
                target_name = "未知参数"
                if dep_id in all_params:
                    target_name = all_params[dep_id].get('名称', dep_id)
                
                dependencies.append({
                    'source': source_name,
                    'source_id': param_id,
                    'target': target_name,
                    'target_id': dep_id
                })
        
        return jsonify(dependencies)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"获取依赖关系时出错: {str(e)}")
        print(f"错误详情: {error_details}")
        return jsonify({'error': f'获取依赖关系时出错: {str(e)}'}), 500

# API: 获取特定参数的详细信息
@app.route('/api/parameter_details/<param_id>')
def get_parameter_details(param_id):
    if not session.get('file_path'):
        return jsonify({'error': '找不到已分析的文件'}), 404
    
    try:
        file_path = session['file_path']
        print(f"正在获取参数详情的文件: {file_path}")
        
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            # 尝试回退到原始文件
            if session.get('original_file_path') and os.path.exists(session['original_file_path']):
                file_path = session['original_file_path']
                print(f"回退到原始文件: {file_path}")
            else:
                return jsonify({'error': f'文件不存在: {file_path}'}), 404
            
        # 重新加载Excel工作簿
        wb = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        
        # 收集参数信息
        all_params, formula_dependencies = excel_analyzer.collect_params_and_dependencies(wb_formulas, wb, {})
        
        if param_id not in all_params:
            return jsonify({'error': '找不到指定的参数'}), 404
        
        param_info = all_params[param_id]
        
        # 将set类型转换为list类型，以便于JSON序列化
        if "依赖" in param_info and isinstance(param_info["依赖"], set):
            param_info["依赖"] = list(param_info["依赖"])
        if "依赖描述" in param_info and isinstance(param_info["依赖描述"], set):
            param_info["依赖描述"] = list(param_info["依赖描述"])
        
        # 获取依赖链
        dependency_chain = get_dependency_chain(param_id, all_params, formula_dependencies)
        
        # 确保参数信息完整
        name = param_info.get('名称', param_id)
        value = param_info.get('值', 0)
        unit = param_info.get('单位', '')
        formula = param_info.get('公式', '')
        formula_description = param_info.get('公式描述', '')
        dependencies = param_info.get('依赖', [])
        dependency_names = param_info.get('依赖描述', [])
        has_circular_dependency = param_info.get('有循环依赖', False)
        
        # 构建详细信息
        details = {
            'id': param_id,
            'name': name,
            'value': value,
            'unit': unit,
            'formula': formula,
            'formula_description': formula_description,
            'dependencies': dependencies,
            'dependency_names': dependency_names,
            'dependency_chain': dependency_chain,
            'has_circular_dependency': has_circular_dependency
        }
        
        return jsonify(details)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"获取参数详细信息时出错: {str(e)}")
        print(f"错误详情: {error_details}")
        return jsonify({'error': f'获取参数详细信息时出错: {str(e)}'}), 500

# 递归获取依赖链
def get_dependency_chain(param_id, all_params, formula_dependencies, visited=None):
    """
    递归获取参数依赖链，添加循环检测以避免无限递归
    
    Args:
        param_id: 当前参数ID
        all_params: 所有参数信息
        formula_dependencies: 参数依赖关系
        visited: 已访问的参数ID集合，用于检测循环依赖
    """
    # 初始化已访问集合（仅在顶层调用时）
    if visited is None:
        visited = set()
    
    # 检查formula_dependencies是否为有效字典
    if not isinstance(formula_dependencies, dict):
        print(f"警告: formula_dependencies不是字典类型: {type(formula_dependencies)}")
        return []
    
    # 如果参数不在依赖关系中或已被访问（循环依赖），返回空列表
    if param_id not in formula_dependencies or param_id in visited:
        return []
    
    # 检查依赖项是否为有效集合或列表
    deps = formula_dependencies.get(param_id, [])
    if not deps:
        return []
        
    # 确保deps是列表类型
    if isinstance(deps, set):
        deps = list(deps)
    elif not isinstance(deps, list):
        print(f"警告: 依赖项不是集合或列表类型: {type(deps)}")
        return []
    
    # 标记当前参数为已访问
    visited.add(param_id)
    
    chain = []
    for dep_id in deps:
        if dep_id in all_params:
            # 检查是否形成循环
            is_cycle = dep_id in visited
            
            # 安全获取参数信息
            dep_name = all_params[dep_id].get('名称', dep_id)
            dep_value = all_params[dep_id].get('值', 0)
            dep_unit = all_params[dep_id].get('单位', '')
            
            dep_info = {
                'id': dep_id,
                'name': dep_name,
                'value': dep_value,
                'unit': dep_unit,
                'is_cycle': is_cycle  # 添加循环标记
            }
            
            # 只有非循环依赖才继续递归
            if not is_cycle:
                # 创建visited的副本进行递归，避免跨分支影响
                dep_info['children'] = get_dependency_chain(dep_id, all_params, formula_dependencies, visited.copy())
            else:
                dep_info['children'] = []  # 循环依赖不再展开
                
            chain.append(dep_info)
    
    return chain

# API: 计算参数值
@app.route('/api/calculate', methods=['POST'])
def calculate_parameters():
    if not session.get('file_path'):
        return jsonify({'error': '找不到已分析的文件'}), 404
    
    try:
        # 获取输入参数值
        input_values = request.json
        file_path = session['file_path']
        print(f"正在进行参数计算的文件: {file_path}")
        
        if not os.path.exists(file_path):
            print(f"文件不存在: {file_path}")
            # 尝试回退到原始文件
            if session.get('original_file_path') and os.path.exists(session['original_file_path']):
                file_path = session['original_file_path']
                print(f"回退到原始文件: {file_path}")
            else:
                return jsonify({'error': f'文件不存在: {file_path}'}), 404
        
        # 重新加载Excel工作簿（仅用于分析，不进行计算）
        wb = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        
        # 收集参数信息和依赖关系
        all_params, formula_dependencies = excel_analyzer.collect_params_and_dependencies(wb_formulas, wb, {})
        
        # 检查并处理输入值
        if not input_values or not isinstance(input_values, dict):
            print("警告: 输入值格式不正确")
            input_values = {}
        
        # 更新输入参数值
        for param_id, value in input_values.items():
            if param_id in all_params:
                try:
                    # 尝试转换为数值类型，但保留字符串格式
                    if isinstance(value, str) and (":" in value):
                        # 对于包含冒号的字符串，保持原始格式
                        all_params[param_id]['值'] = value
                    else:
                        # 尝试转换为数值
                        all_params[param_id]['值'] = float(value)
                except (ValueError, TypeError):
                    print(f"警告: 无法将输入值转换为数字: param_id={param_id}, value={value}")
                    # 对于无法转换的值，保持原始格式
                    all_params[param_id]['值'] = value
        
        # 按依赖关系顺序计算所有参数值
        sorted_params = topological_sort(all_params, formula_dependencies)
        
        # 使用xlwings进行Excel直接计算（不再使用formulas库）
        calculated_values = calculate_values(sorted_params, all_params, formula_dependencies)
        
        return jsonify({'calculated_values': calculated_values})
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"计算参数值时出错: {str(e)}")
        print(f"错误详情: {error_details}")
        return jsonify({'error': f'计算参数值时出错: {str(e)}'}), 500

# 拓扑排序 - 确保按依赖顺序计算参数
def topological_sort(all_params, formula_dependencies):
    """
    使用Kahn算法进行拓扑排序，能够处理循环依赖的情况
    """
    # 检查参数和依赖关系是否有效
    if not isinstance(all_params, dict) or not isinstance(formula_dependencies, dict):
        print(f"警告: 无效的数据结构 - all_params: {type(all_params)}, formula_dependencies: {type(formula_dependencies)}")
        # 直接返回参数ID列表，无法排序
        return list(all_params.keys())
        
    # 创建入度表
    in_degree = {param_id: 0 for param_id in all_params}
    
    # 计算每个节点的入度
    for param_id, deps in formula_dependencies.items():
        # 确保deps是集合或列表
        if not deps:
            continue
            
        deps_list = list(deps) if isinstance(deps, set) else deps
        if not isinstance(deps_list, list):
            print(f"警告: 依赖项不是有效的集合或列表: {type(deps)}")
            continue
            
        for dep_id in deps_list:
            if dep_id in all_params:  # 只考虑存在的参数
                in_degree[dep_id] = in_degree.get(dep_id, 0) + 1
    
    # 获取所有入度为0的节点（没有依赖或只依赖外部参数）
    zero_in_degree = [param_id for param_id in all_params if in_degree.get(param_id, 0) == 0]
    
    # 结果列表
    result = []
    
    # 循环直到没有入度为0的节点
    while zero_in_degree:
        # 移除一个入度为0的节点
        current = zero_in_degree.pop(0)
        result.append(current)
        
        # 减少所有依赖该节点的节点的入度
        if current in formula_dependencies:
            deps = formula_dependencies[current]
            deps_list = list(deps) if isinstance(deps, set) else deps
            
            if not isinstance(deps_list, list):
                continue
                
            for neighbor in deps_list:
                if neighbor in in_degree:
                    in_degree[neighbor] -= 1
                    
                    # 如果入度变为0，则加入队列
                    if in_degree[neighbor] == 0:
                        zero_in_degree.append(neighbor)
    
    # 检查是否有循环依赖
    remaining = [param_id for param_id in all_params if param_id not in result]
    if remaining:
        print(f"警告: 检测到循环依赖，这些参数将被添加到排序尾部: {remaining}")
        result.extend(remaining)  # 将剩余的节点添加到结果末尾
    
    return result

# 根据参数排序计算值
def calculate_values(sorted_params, all_params, formula_dependencies):
    calculated_values = {}
    
    try:
        # 使用xlwings直接与Excel交互计算
        file_path = session.get('file_path')
        if not file_path or not os.path.exists(file_path):
            # 尝试回退到原始文件
            file_path = session.get('original_file_path')
            if not file_path or not os.path.exists(file_path):
                print("找不到Excel文件，无法使用xlwings进行计算")
                raise FileNotFoundError("找不到Excel文件")
        
        print(f"使用xlwings打开Excel文件: {file_path}")
        
        # 获取输入参数和依赖信息
        input_params, output_params, intermediate_params, independent_params = excel_analyzer.categorize_parameters(all_params, formula_dependencies)
        
        # 使用xlwings打开Excel应用程序和工作簿
        try:
            # 在可见模式下启动Excel（调试时使用，部署时可设为False）
            app = xw.App(visible=False)
            wb = app.books.open(file_path)
            
            # 遍历所有参数，处理输入参数（更新Excel中的值）
            print("更新Excel中的输入参数值...")
            for param_id in input_params:
                param_info = all_params.get(param_id, {})
                if not param_info:
                    continue
                
                # 获取参数位置信息
                sheet_name = param_info.get('工作表', '')
                row = param_info.get('行', 0)
                value = param_info.get('值', 0)
                
                if sheet_name and row > 0:
                    try:
                        # 在Excel中更新输入值（默认第3列为值列）
                        sheet = wb.sheets[sheet_name]
                        cell = sheet.cells(row, 3)  # 第3列是值列
                        print(f"更新参数 {param_id} 在单元格 {sheet_name}!C{row} 的值为 {value}")
                        cell.value = value
                    except Exception as e:
                        print(f"更新Excel中的输入参数 {param_id} 时出错: {str(e)}")
            
            # 等待Excel重新计算
            print("等待Excel重新计算...")
            wb.app.calculate()
            
            # 读取计算后的输出和中间参数值
            print("读取Excel中的计算结果...")
            result_params = intermediate_params.union(output_params)
            
            for param_id in result_params:
                param_info = all_params.get(param_id, {})
                if not param_info:
                    continue
                
                # 获取参数位置信息
                sheet_name = param_info.get('工作表', '')
                row = param_info.get('行', 0)
                name = param_info.get('名称', param_id)
                unit = param_info.get('单位', '')
                formula = param_info.get('公式', '')
                
                if sheet_name and row > 0:
                    try:
                        # 从Excel中读取计算后的值
                        sheet = wb.sheets[sheet_name]
                        cell = sheet.cells(row, 3)  # 第3列是值列
                        calculated_value = cell.value
                        print(f"读取参数 {param_id} 在单元格 {sheet_name}!C{row} 的计算结果: {calculated_value}")
                        
                        # 特殊处理坡度表示格式和其他字符串类型结果
                        if isinstance(calculated_value, str):
                            # 已经是字符串格式，保持不变
                            formatted_value = calculated_value
                        elif isinstance(formula, str) and ('&' in formula or 'CONCATENATE' in formula.upper()):
                            # 检查公式包含字符串连接操作，但结果可能被转为数值
                            # 尝试根据公式特征自行格式化结果
                            if "1:" in formula or "1：" in formula:
                                # 对于坡度表示，添加"1:"前缀
                                formatted_value = f"1:{calculated_value}" if calculated_value else "1:0"
                            else:
                                # 其他情况仍使用原始值
                                formatted_value = calculated_value
                        else:
                            # 使用原始值
                            formatted_value = calculated_value
                        
                        calculated_values[param_id] = {
                            'id': param_id,
                            'name': name,
                            'value': formatted_value,
                            'unit': unit
                        }
                    except Exception as e:
                        print(f"读取Excel中的参数 {param_id} 计算结果时出错: {str(e)}")
                        # 使用原始值作为备选
                        calculated_values[param_id] = {
                            'id': param_id,
                            'name': name,
                            'value': param_info.get('值', 0),
                            'unit': unit,
                            'error': f"读取错误: {str(e)}"
                        }
            
            # 保存输入参数的计算结果
            for param_id in input_params:
                param_info = all_params.get(param_id, {})
                if not param_info:
                    continue
                
                name = param_info.get('名称', param_id)
                value = param_info.get('值', 0)
                unit = param_info.get('单位', '')
                
                calculated_values[param_id] = {
                    'id': param_id,
                    'name': name,
                    'value': value,
                    'unit': unit
                }
                
        except Exception as e:
            print(f"xlwings操作Excel时出错: {str(e)}")
            raise
        finally:
            # 清理资源，确保关闭Excel
            try:
                if 'wb' in locals() and wb:
                    wb.close()
                if 'app' in locals() and app:
                    app.quit()
            except Exception as e:
                print(f"关闭Excel资源时出错: {str(e)}")
        
        return calculated_values
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"使用xlwings计算时出错: {str(e)}")
        print(f"错误详情: {error_details}")
        
        # 如果xlwings方法失败，提供合理的错误信息并返回
        for param_id in sorted_params:
            param_info = all_params.get(param_id, {})
            if not param_info:
                continue
            
            name = param_info.get('名称', param_id)
            value = param_info.get('值', 0)
            unit = param_info.get('单位', '')
            
            calculated_values[param_id] = {
                'id': param_id,
                'name': name,
                'value': value,
                'unit': unit,
                'error': f"计算错误: {str(e)}"
            }
    
    return calculated_values

if __name__ == '__main__':
    app.run(debug=True) 