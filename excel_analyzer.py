import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
from openpyxl.styles import PatternFill, Font

def analyze_excel(file_path):
    """分析Excel文件中的参数、公式和依赖关系，并生成优化后的Excel文件"""
    # 加载Excel工作簿（一个包含公式，一个包含计算后的值）
    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    # 收集所有参数和依赖关系
    all_params = {}
    formula_dependencies = {}
    
    # 存储重名参数信息
    duplicate_params = collect_duplicate_params(wb)
    
    # 收集参数和依赖关系
    all_params, formula_dependencies = collect_params_and_dependencies(wb, wb_data, duplicate_params)
    
    # 处理重名参数、依赖关系和参数分类
    param_replacements, different_value_groups, optimized_dependencies, renamed_params = process_parameters(all_params, formula_dependencies)
    
    # 生成优化后的Excel文件
    optimized_excel_path = generate_optimized_excel(file_path, all_params, param_replacements, different_value_groups, formula_dependencies)
    
    if optimized_excel_path:
        print(f"优化后的Excel文件已保存至: {optimized_excel_path}")
    else:
        print("无法优化Excel文件，将使用原始文件。")
    
    return all_params

def collect_duplicate_params(wb):
    """收集Excel中的重名参数"""
    duplicate_params = {}  # 格式: {参数名: [位置1, 位置2, ...]}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        name_col = 1
        
        # 遍历所有行（从第二行开始，第一行是表头）
        for row in range(2, ws.max_row + 1):
            param_name = ws.cell(row=row, column=name_col).value
            if not param_name:
                continue  # 跳过没有参数名的行
            
            # 记录重名参数
            if param_name in duplicate_params:
                duplicate_params[param_name].append((sheet, row))
            else:
                duplicate_params[param_name] = [(sheet, row)]
    
    return duplicate_params

def collect_params_and_dependencies(wb, wb_data, duplicate_params):
    """收集所有参数和依赖关系"""
    all_params = {}
    formula_dependencies = {}
    
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws_data = wb_data[sheet]
            
            name_col = 1
            unit_col = 2
            value_col = 3
            
            # 检查工作表结构
            if ws.max_row < 2 or ws.max_column < 3:
                print(f"警告: 工作表 {sheet} 结构不符合要求")
                continue
            
            # 创建单元格到参数名的映射
            cell_to_param_map = {}
            
            # 先建立单元格地址到参数标识符的映射
            for row in range(2, ws.max_row + 1):
                param_name = ws.cell(row=row, column=name_col).value
                if param_name:
                    # 检查是否为重名参数
                    is_duplicate = len(duplicate_params.get(param_name, [])) > 1
                    
                    # 为重名参数创建唯一标识符
                    param_id = f"{param_name}_{sheet}_r{row}" if is_duplicate else param_name
                    
                    cell_addr = f"{get_column_letter(name_col)}{row}"
                    cell_to_param_map[cell_addr] = param_id
            
            # 遍历所有行收集参数信息和依赖关系
            for row in range(2, ws.max_row + 1):
                param_name = ws.cell(row=row, column=name_col).value
                if not param_name:
                    continue  # 跳过没有参数名的行
                
                # 检查是否为重名参数
                is_duplicate = len(duplicate_params.get(param_name, [])) > 1
                
                # 为重名参数创建唯一标识符
                param_id = f"{param_name}_{sheet}_r{row}" if is_duplicate else param_name
                
                param_unit = ws.cell(row=row, column=unit_col).value
                
                # 存储参数信息
                param_info = {
                    "名称": param_name,
                    "标识符": param_id,
                    "单位": param_unit if param_unit else "",
                    "工作表": sheet,
                    "行": row,
                    "值": None,
                    "公式": "",
                    "公式描述": "",
                    "依赖": set(),
                    "依赖描述": set(),
                    "是否继承": False,
                    "有循环依赖": False
                }
                
                # 获取数值
                value_cell = ws.cell(row=row, column=value_col)
                calculated_cell = ws_data.cell(row=row, column=value_col)
                
                # 检查是否为公式
                if value_cell.data_type == 'f':
                    original_formula = str(value_cell.value)
                    
                    # 去除公式前的等号
                    if original_formula.startswith('='):
                        original_formula = original_formula[1:]
                    
                    calculated_value = calculated_cell.value
                    
                    param_info["值"] = calculated_value
                    param_info["公式"] = original_formula
                    
                    # 分析公式中的依赖关系
                    try:
                        direct_refs = re.findall(r'([A-Z]+)([0-9]+)', original_formula)
                        range_refs = re.findall(r'([A-Z]+)([0-9]+):([A-Z]+)([0-9]+)', original_formula)
                        
                        human_readable_formula = original_formula
                        
                        # 处理直接引用分析
                        for col_letter, row_num in direct_refs:
                            cell_addr = f"{col_letter}{row_num}"
                            ref_row = int(row_num)
                            
                            # 如果引用了另一个参数（行）
                            if ref_row != row and ref_row >= 2:  # 不是自己且不是表头
                                ref_param_name = ws.cell(row=ref_row, column=name_col).value
                                if ref_param_name:
                                    # 检查被引用参数是否为重名参数
                                    is_ref_duplicate = len(duplicate_params.get(ref_param_name, [])) > 1
                                    ref_param_id = f"{ref_param_name}_{sheet}_r{ref_row}" if is_ref_duplicate else ref_param_name
                                    
                                    param_info["依赖"].add(ref_param_id)
                                    param_info["依赖描述"].add(ref_param_name)
                                    
                                    # 存储依赖关系
                                    if param_id not in formula_dependencies:
                                        formula_dependencies[param_id] = set()
                                    formula_dependencies[param_id].add(ref_param_id)
                                    
                                    # 替换公式中的单元格引用为参数名
                                    if cell_addr in human_readable_formula:
                                        human_readable_formula = human_readable_formula.replace(cell_addr, ref_param_name)
                        
                        # 处理范围引用
                        for start_col_letter, start_row, end_col_letter, end_row in range_refs:
                            start_row = int(start_row)
                            end_row = int(end_row)
                            
                            for r in range(start_row, end_row + 1):
                                if r != row and r >= 2:  # 不是自己且不是表头
                                    ref_param_name = ws.cell(row=r, column=name_col).value
                                    if ref_param_name:
                                        # 检查被引用参数是否为重名参数
                                        is_ref_duplicate = len(duplicate_params.get(ref_param_name, [])) > 1
                                        ref_param_id = f"{ref_param_name}_{sheet}_r{r}" if is_ref_duplicate else ref_param_name
                                        
                                        param_info["依赖"].add(ref_param_id)
                                        param_info["依赖描述"].add(ref_param_name)
                                        
                                        # 存储依赖关系
                                        if param_id not in formula_dependencies:
                                            formula_dependencies[param_id] = set()
                                        formula_dependencies[param_id].add(ref_param_id)
                        
                        # 更新公式描述
                        param_info["公式描述"] = human_readable_formula
                    except Exception as e:
                        print(f"分析公式时出错: {str(e)}")
                        param_info["公式描述"] = "公式分析错误: " + original_formula
                else:
                    # 非公式值
                    param_info["值"] = value_cell.value
                
                # 将参数信息添加到总字典中
                all_params[param_id] = param_info
        
        # 检测循环依赖
        circular_dependencies, cycle_paths = detect_circular_dependencies(formula_dependencies)
        
        # 在参数信息中标记循环依赖
        for param_id in circular_dependencies:
            if param_id in all_params:
                all_params[param_id]["有循环依赖"] = True
        
        return all_params, formula_dependencies
    except Exception as e:
        print(f"收集参数和依赖关系时出错: {str(e)}")
        return all_params, formula_dependencies

def detect_circular_dependencies(formula_dependencies):
    """简化的循环依赖检测算法"""
    circular_params = set()  # 存储有循环依赖的参数ID
    cycle_paths = []  # 存储发现的循环路径
    
    # 使用简化的DFS算法
    for start_node in formula_dependencies:
        visited = set()
        stack = [(start_node, [start_node])]
        
        while stack:
            node, path = stack.pop()
            
            # 获取当前节点的依赖
            deps = formula_dependencies.get(node, set())
            for dep in deps:
                # 如果依赖已在当前路径中，发现循环
                if dep in path:
                    cycle = path[path.index(dep):] + [dep]
                    cycle_paths.append(cycle)
                    circular_params.update(cycle)
                # 否则继续DFS
                elif dep in formula_dependencies and dep not in visited:
                    visited.add(dep)
                    stack.append((dep, path + [dep]))
    
    # 去除重复循环路径
    unique_paths = []
    seen = set()
    
    for cycle in cycle_paths:
        # 标准化循环路径以检测重复
        cycle_str = '->'.join(sorted(cycle))
        if cycle_str not in seen:
            seen.add(cycle_str)
            unique_paths.append(cycle)
    
    return circular_params, unique_paths

def process_parameters(all_params, formula_dependencies):
    """处理参数：先处理同名同值参数，再处理同名不同值参数"""
    # 重复参数处理结果
    param_replacements = {}  # 被替换的参数: 源参数
    different_value_groups = {}  # 同名不同值的参数分组
    renamed_params = {}  # 需要重命名的参数
    
    # 建立参数组
    param_groups = {}  # {参数名: {参数ID: 参数信息}}
    for param_id, param_info in all_params.items():
        param_name = param_info["名称"]
        if param_name not in param_groups:
            param_groups[param_name] = {}
        param_groups[param_name][param_id] = param_info
    
    # 第一步：处理同名同值参数
    for param_name, param_group in param_groups.items():
        if len(param_group) <= 1:
            continue  # 没有重名，跳过
        
        # 按值分组
        value_groups = {}
        for param_id, param_info in param_group.items():
            value = param_info["值"]
            
            if value not in value_groups:
                value_groups[value] = []
            value_groups[value].append((param_id, param_info))
        
        # 对每个值组合内的多个参数进行处理（同名同值的情况）
        for value, params in value_groups.items():
            if len(params) <= 1:
                continue  # 该值只有一个参数，不需处理
            
            # 选择源参数（保留一个，其余的将被删除）
            source_param_id = None
            
            # 对参数按照优先级排序: 无公式无依赖 > 有公式/有依赖
            params_with_priority = []
            for i, (param_id, param_info) in enumerate(params):
                has_formula = bool(param_info.get("公式", ""))
                has_dependency = bool(param_info.get("依赖", set()))
                
                # 计算优先级: 无公式无依赖的优先级最高
                priority = 0 if (not has_formula and not has_dependency) else 1
                params_with_priority.append((priority, i, param_id, param_info))
            
            # 按优先级排序
            params_with_priority.sort()
            
            # 选择优先级最高的参数作为源参数
            if params_with_priority:
                _, _, source_param_id, _ = params_with_priority[0]
            
            # 标记除源参数外的所有参数为被替换参数
            for param_id, _ in params:
                if param_id != source_param_id:
                    param_replacements[param_id] = source_param_id
    
    # 应用参数替换，更新all_params
    # 移除被替换的参数，创建剩余参数集合
    surviving_params = {}
    for param_id, param_info in all_params.items():
        if param_id not in param_replacements:
            surviving_params[param_id] = param_info
    
    # 第二步：处理剩余的同名不同值参数
    # 重建参数组，排除已被替换的参数
    remaining_param_groups = {}
    for param_id, param_info in surviving_params.items():
        param_name = param_info["名称"]
        if param_name not in remaining_param_groups:
            remaining_param_groups[param_name] = {}
        remaining_param_groups[param_name][param_id] = param_info
    
    # 处理剩余的同名不同值参数
    for param_name, param_group in remaining_param_groups.items():
        if len(param_group) <= 1:
            continue  # 没有重名，跳过
        
        # 同名不同值的情况，需要进行重命名
        # 按值分组
        value_groups = {}
        for param_id, param_info in param_group.items():
            value = param_info["值"]
            
            if value not in value_groups:
                value_groups[value] = []
            value_groups[value].append((param_id, param_info))
        
        # 为每个不同值的组合创建命名
        counter = 1
        for value, params in sorted(value_groups.items(), key=lambda x: (x[0] is None, x[0])):
            group_name = param_name if counter == 1 else f"{param_name}_{counter}"
            counter += 1
            
            different_value_groups.setdefault(param_name, {})[value] = [param_id for param_id, _ in params]
            
            # 重命名非第一组参数
            if group_name != param_name:
                for param_id, param_info in params:
                    renamed_params[param_id] = group_name
                    all_params[param_id]["名称"] = group_name
                    
                    # 更新公式描述中的参数名
                    if "公式描述" in param_info and param_info["公式描述"]:
                        formula_desc = param_info["公式描述"]
                        all_params[param_id]["公式描述"] = formula_desc.replace(param_name, group_name)
    
    # 优化依赖关系，替换重复参数引用
    optimized_dependencies = {}
    for param_id, deps in formula_dependencies.items():
        # 跳过被替换的参数
        if param_id in param_replacements:
            continue
        
        # 替换依赖中的重复参数引用
        new_deps = set()
        for dep_id in deps:
            if dep_id in param_replacements:
                # 将依赖的被替换参数替换为源参数
                new_deps.add(param_replacements[dep_id])
            else:
                new_deps.add(dep_id)
        
        optimized_dependencies[param_id] = new_deps
    
    return param_replacements, different_value_groups, optimized_dependencies, renamed_params

def generate_optimized_excel(file_path, all_params, param_replacements, different_value_groups, formula_dependencies):
    """生成优化后的Excel文件，处理同名同值参数"""
    output_path = os.path.splitext(file_path)[0] + "_optimized.xlsx"
    
    try:
        # 加载原始Excel工作簿
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # 创建参数ID到位置的映射
        param_id_to_location = {}
        location_to_param_id = {}
        
        for param_id, param_info in all_params.items():
            sheet_name = param_info.get("工作表")
            row = param_info.get("行")
            param_id_to_location[param_id] = (sheet_name, row)
            location_to_param_id[(sheet_name, row)] = param_id
        
        # 确定参数类型
        input_params, output_params, intermediate_params, _ = categorize_parameters(all_params, formula_dependencies)
        
        # 定义颜色填充方案
        fills = {
            'input': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # 浅蓝色-输入参数
            'output': PatternFill(start_color="F08080", end_color="F08080", fill_type="solid"),  # 浅红色-输出参数
            'intermediate': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # 浅绿色-中间参数
            'circular': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # 黄色-循环依赖
            'replaced': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")   # 灰色-被替换参数
        }
        
        # 更新Excel中的参数名称并添加颜色标记
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            name_col = 1
            
            # 找出每个工作表的实际最大列，避免出现空列
            actual_max_col = 0
            for col in range(1, ws.max_column + 1):
                for row in range(1, ws.max_row + 1):
                    if ws.cell(row=row, column=col).value is not None:
                        actual_max_col = max(actual_max_col, col)
            
            # 设置依赖关系和公式列的位置（紧接实际数据列）
            dependency_col = actual_max_col + 1
            formula_col = actual_max_col + 2
            
            # 更新参数名称和添加颜色标记
            for row in range(2, ws.max_row + 1):
                param_id = location_to_param_id.get((sheet_name, row))
                if not param_id or param_id not in all_params:
                    continue
                
                # 更新参数名称
                param_name = all_params[param_id].get("名称")
                if param_name:
                    ws.cell(row=row, column=name_col).value = param_name
                
                # 添加颜色标记
                fill = None
                
                # 对被替换的参数进行灰色标记（将被删除）
                if param_id in param_replacements:
                    fill = fills['replaced']
                    source_id = param_replacements[param_id]
                    source_name = all_params[source_id]["名称"]
                    # 在备注中添加被替换信息
                    cell = ws.cell(row=row, column=name_col)
                    if not cell.comment:
                        comment = f"此参数将被删除，使用 {source_name} 替代"
                        cell.comment = openpyxl.comments.Comment(comment, "Excel分析工具")
                # 对其他类型参数进行颜色标记
                elif all_params[param_id].get("有循环依赖", False):
                    fill = fills['circular']  # 循环依赖
                elif param_id in input_params:
                    fill = fills['input']  # 输入参数
                elif param_id in intermediate_params:
                    fill = fills['intermediate']  # 中间参数
                elif param_id in output_params:
                    fill = fills['output']  # 输出参数
                
                # 应用颜色
                if fill:
                    for col in range(1, 4):  # 只为前三列添加颜色（名称、单位、值）
                        ws.cell(row=row, column=col).fill = fill
            
            # 添加依赖关系和公式列
            ws.cell(row=1, column=dependency_col).value = "依赖关系"
            ws.cell(row=1, column=formula_col).value = "公式"
            ws.cell(row=1, column=dependency_col).font = Font(bold=True)
            ws.cell(row=1, column=formula_col).font = Font(bold=True)
            
            # 填充依赖关系和公式
            for row in range(2, ws.max_row + 1):
                param_id = location_to_param_id.get((sheet_name, row))
                if not param_id or param_id not in all_params:
                    continue
                
                param_info = all_params[param_id]
                
                # 填充依赖关系
                dependencies = list(param_info.get("依赖描述", set()))
                if dependencies:
                    ws.cell(row=row, column=dependency_col).value = ", ".join(dependencies)
                
                # 填充公式
                formula_desc = param_info.get("公式描述", "")
                if formula_desc:
                    ws.cell(row=row, column=formula_col).value = formula_desc
                    
                # 对于被替换的参数，添加替换信息
                if param_id in param_replacements:
                    source_id = param_replacements[param_id]
                    source_info = all_params[source_id]
                    replacement_note = f"将被 {source_info['名称']} (工作表:{source_info['工作表']}, 行:{source_info['行']}) 替代"
                    ws.cell(row=row, column=formula_col+1).value = replacement_note
        
        # 删除被替换的参数行
        row_shifts = delete_replaced_rows(wb, all_params, param_replacements, location_to_param_id)
        
        # 修复公式引用
        fix_formula_references(wb, param_id_to_location, param_replacements, row_shifts)
        
        # 保存优化后的Excel
        wb.save(output_path)
        return output_path
    
    except Exception as e:
        print(f"生成优化后的Excel时出错: {str(e)}")
        return None

def categorize_parameters(all_params, formula_dependencies):
    """对参数进行分类"""
    # 找出所有参与依赖关系的参数
    all_dependent_params = set()  # 被依赖的参数
    all_dependency_params = set()  # 依赖其他参数的参数
    circular_params = set()  # 循环依赖的参数
    
    # 找出循环依赖参数
    for param_id, param_info in all_params.items():
        if param_info.get("有循环依赖", False):
            circular_params.add(param_id)
    
    # 收集依赖关系
    for param, deps in formula_dependencies.items():
        all_dependency_params.add(param)
        for dep in deps:
            all_dependent_params.add(dep)
    
    # 分类参数
    input_params = (all_dependent_params - all_dependency_params) - circular_params
    output_params = (all_dependency_params - all_dependent_params) - circular_params
    intermediate_params = (all_dependent_params.intersection(all_dependency_params)) | circular_params
    
    # 独立参数
    independent_params = set()
    for param_id in all_params:
        if param_id not in input_params and param_id not in output_params and param_id not in intermediate_params:
            independent_params.add(param_id)
    
    return input_params, output_params, intermediate_params, independent_params

def delete_replaced_rows(wb, all_params, param_replacements, location_to_param_id):
    """删除被替换的参数行，并处理相关行号变化"""
    row_shifts = {}  # {(sheet_name, original_row): shift_count}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_col = 1
        
        # 收集要删除的行
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            param_id = location_to_param_id.get((sheet_name, row))
            if param_id and param_id in param_replacements:
                # 这一行将被删除，因为它是被替换的参数
                rows_to_delete.append(row)
        
        if not rows_to_delete:
            continue
        
        # 从后往前删除行，避免索引变化影响
        rows_to_delete.sort(reverse=True)
        shift_count = 0
        
        for row in rows_to_delete:
            # 更新此行之后所有行的位移信息
            for r in range(row + 1, ws.max_row + 2):
                row_shifts[(sheet_name, r)] = shift_count
            
            # 删除行
            ws.delete_rows(row)
            shift_count += 1
        
        # 更新最终的位移计数
        if rows_to_delete:
            min_deleted_row = min(rows_to_delete)
            for r in range(min_deleted_row, ws.max_row + 2):
                row_shifts[(sheet_name, r)] = shift_count
    
    return row_shifts

def fix_formula_references(wb, param_id_to_location, param_replacements, row_shifts=None):
    """修复公式引用，将被替换参数的引用替换为源参数引用"""
    if row_shifts is None:
        row_shifts = {}
    
    # 创建源参数位置查找表，考虑行位移
    source_locations = {}
    for dependent_id, source_id in param_replacements.items():
        if dependent_id in param_id_to_location and source_id in param_id_to_location:
            dependent_loc = param_id_to_location[dependent_id]
            source_loc = param_id_to_location[source_id]
            source_sheet, source_row = source_loc
            
            # 应用行位移调整
            adjusted_row = source_row - row_shifts.get((source_sheet, source_row), 0)
            source_locations[dependent_id] = (source_sheet, adjusted_row)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                if cell.data_type == 'f':
                    original_formula = str(cell.value)
                    new_formula = original_formula
                    has_changes = False
                    
                    # 1. 替换从属参数引用为源参数引用
                    for dependent_id, source_id in param_replacements.items():
                        if dependent_id not in param_id_to_location or source_id not in param_id_to_location:
                            continue
                            
                        dependent_sheet, dependent_row = param_id_to_location[dependent_id]
                        source_sheet, adjusted_row = source_locations.get(dependent_id, param_id_to_location[source_id])
                        
                        # 替换带工作表名称的引用
                        for col_letter in ['A', 'C']:  # 参数名称列和值列
                            dep_addr = f"{dependent_sheet}!{col_letter}{dependent_row}"
                            src_addr = f"{source_sheet}!{col_letter}{adjusted_row}"
                            
                            if dep_addr in new_formula:
                                new_formula = new_formula.replace(dep_addr, src_addr)
                                has_changes = True
                        
                        # 替换同一工作表内的简单引用
                        if sheet_name == dependent_sheet:
                            for col_letter in ['A', 'C']:
                                dep_addr = f"{col_letter}{dependent_row}"
                                
                                if source_sheet == sheet_name:
                                    src_addr = f"{col_letter}{adjusted_row}"
                                else:
                                    src_addr = f"{source_sheet}!{col_letter}{adjusted_row}"
                                    
                                if dep_addr in new_formula:
                                    new_formula = new_formula.replace(dep_addr, src_addr)
                                    has_changes = True
                    
                    # 2. 处理行位移引起的单元格引用变化
                    if row_shifts:
                        # 处理直接引用
                        cell_refs = re.findall(r'([A-Za-z]+)([0-9]+)', new_formula)
                        for col_letter, row_num in cell_refs:
                            ref_row = int(row_num)
                            shift = row_shifts.get((sheet_name, ref_row), 0)
                            
                            if shift > 0:
                                new_row = ref_row - shift
                                old_ref = f"{col_letter}{row_num}"
                                new_ref = f"{col_letter}{new_row}"
                                new_formula = new_formula.replace(old_ref, new_ref)
                                has_changes = True
                        
                        # 处理范围引用
                        range_refs = re.findall(r'([A-Za-z]+)([0-9]+):([A-Za-z]+)([0-9]+)', new_formula)
                        for start_col, start_row, end_col, end_row in range_refs:
                            # 处理范围的行位移
                            start_shift = row_shifts.get((sheet_name, int(start_row)), 0)
                            end_shift = row_shifts.get((sheet_name, int(end_row)), 0)
                            
                            if start_shift > 0 or end_shift > 0:
                                new_start = int(start_row) - start_shift
                                new_end = int(end_row) - end_shift
                                
                                old_range = f"{start_col}{start_row}:{end_col}{end_row}"
                                new_range = f"{start_col}{new_start}:{end_col}{new_end}"
                                
                                new_formula = new_formula.replace(old_range, new_range)
                                has_changes = True
                    
                    # 更新单元格公式
                    if has_changes:
                        try:
                            cell.value = new_formula
                        except Exception as e:
                            print(f"无法更新公式: {str(e)}")

def main():
    print("Excel参数分析工具")
    print("=================")
    file_path = input("请输入Excel文件的路径: ")
    try:
        analyze_excel(file_path)
    except Exception as e:
        print(f"发生错误: {str(e)}")

if __name__ == "__main__":
    main() 